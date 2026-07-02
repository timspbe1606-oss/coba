streamlit_app_code_fpdf2 = '''
import streamlit as st
import pandas as pd
import openpyxl
from fpdf import FPDF # Now imports from fpdf2, which replaces original fpdf

# Load the data - crucial for the Streamlit app to be standalone
try:
    df = pd.read_excel('clean_data_with_sls_match.xlsx')
except FileNotFoundError:
    st.error("Error: 'clean_data_with_sls_match.xlsx' not found. Please ensure it's in the same directory as app.py or the specified path.")
    st.stop()

st.set_page_config(page_title="Usaha Online Musi Banyuasin", layout="wide")
st.title("Usaha Online Musi Banyuasin")

st.markdown("---")

st.sidebar.header("Filter Data")

required_columns = ['nmkab', 'nmkec', 'nmdesa', 'nmsls', 'latitude', 'longitude']
for col in required_columns:
    if col not in df.columns:
        st.error(f"Required column '{col}' not found in the Excel data. Please check your data file.")
        st.stop()

kabupaten_options = ['All'] + sorted(df['nmkab'].astype(str).unique().tolist())
kecamatan_options = ['All'] + sorted(df['nmkec'].astype(str).unique().tolist())
desa_options = ['All'] + sorted(df['nmdesa'].astype(str).unique().tolist())
sls_options = ['All'] + sorted(df['nmsls'].astype(str).unique().tolist())

selected_kabupaten = st.sidebar.selectbox("Kabupaten", kabupaten_options)
selected_kecamatan = st.sidebar.selectbox("Kecamatan", kecamatan_options)
selected_desa = st.sidebar.selectbox("Desa", desa_options)
selected_sls = st.sidebar.selectbox("SLS", sls_options)

filtered_df = df.copy()

if selected_kabupaten != 'All':
    filtered_df = filtered_df[filtered_df['nmkab'].astype(str) == selected_kabupaten]
if selected_kecamatan != 'All':
    filtered_df = filtered_df[filtered_df['nmkec'].astype(str) == selected_kecamatan]
if selected_desa != 'All':
    filtered_df = filtered_df[filtered_df['nmdesa'].astype(str) == selected_desa]
if selected_sls != 'All':
    filtered_df = filtered_df[filtered_df['nmsls'].astype(str) == selected_sls]

st.subheader("Filtered Data Overview")
st.write(f"Showing {len(filtered_df)} records.")

map_data_available = False
if 'latitude' in filtered_df.columns and 'longitude' in filtered_df.columns:
    map_df = filtered_df.dropna(subset=['latitude', 'longitude']).copy()
    if not map_df.empty:
        map_df['latitude'] = pd.to_numeric(map_df['latitude'], errors='coerce')
        map_df['longitude'] = pd.to_numeric(map_df['longitude'], errors='coerce')
        map_df = map_df.dropna(subset=['latitude', 'longitude'])
        if not map_df.empty:
            map_data_available = True
        else:
            st.warning("No valid numeric Latitude/Longitude data after conversion for mapping.")
    else:
        st.warning("No data with valid Latitude/Longitude found for mapping after filtering.")
else:
    st.warning("Latitude or Longitude columns not found in the data for mapping.")


if map_data_available:
    st.subheader("Location Map")
    st.map(map_df[['latitude', 'longitude']])
else:
    st.info("Map cannot be displayed due to missing, invalid, or non-numeric location data in the filtered dataset.")

st.subheader("Detailed Data Table")

display_columns_map = {
    'nmkec': 'Kecamatan',
    'nmdesa': 'Desa',
    'nmsls': 'SLS',
    'Nama Tempat': 'Nama Tempat',
    'Alamat': 'Alamat',
    'Kontak/No HP': 'Kontak/No HP',
    'URL': 'URL'
}

existing_display_cols_for_data_editor = [col for col in display_columns_map.keys() if col in filtered_df.columns]
display_df_for_editor = filtered_df[existing_display_cols_for_data_editor].rename(columns=display_columns_map)

column_config_dict = {}
if 'URL' in existing_display_cols_for_data_editor:
    column_config_dict['URL'] = st.column_config.LinkColumn(
        "URL Link",
        help="Click to open the URL",
        max_chars=100,
        display_text="Open Link"
    )

if not display_df_for_editor.empty:
    st.data_editor(
        display_df_for_editor,
        column_config=column_config_dict,
        hide_index=True,
        use_container_width=True
    )
else:
    st.warning("No data to display in the detailed table after filtering.")


# --- PDF Download Functionality (using fpdf2 for better Unicode support) ---
def create_pdf_from_dataframe(df_to_pdf):
    pdf = FPDF()
    pdf.add_page()
    # Set font for fpdf2. Arial and Helvetica often have good basic Unicode coverage.
    # For full Unicode support, consider adding a specific TTF font like 'DejaVuSans'.
    pdf.set_font("Arial", size=10)

    pdf.cell(200, 10, txt="Usaha Online Musi Banyuasin Data", ln=True, align='C')
    pdf.ln(10)

    col_names = df_to_pdf.columns.tolist()
    if col_names:
        available_width = pdf.w - 2 * pdf.l_margin
        col_width = available_width / len(col_names)

        for header in col_names:
            # fpdf2 handles Unicode strings directly, no need for encode/decode trick
            pdf.cell(col_width, 10, str(header), border=1, align='C')
        pdf.ln()

        for index, row in df_to_pdf.iterrows():
            for col_data in row:
                pdf.cell(col_width, 10, str(col_data), border=1, align='L')
            pdf.ln()
    else:
        pdf.cell(200, 10, "No data available.", ln=True, align='C')

    return pdf.output(dest='S')

if not filtered_df.empty:
    pdf_output = create_pdf_from_dataframe(filtered_df)
    st.download_button(
        label="Download Filtered Data as PDF",
        data=pdf_output,
        file_name="filtered_data.pdf",
        mime="application/pdf",
        key="download_pdf_button"
    )
else:
    st.info("No data available to download as PDF.")

st.markdown("---")
st.caption("Developed with Streamlit for Usaha Online Musi Banyuasin")
'''

with open('app.py', 'w', encoding='utf-8') as f:
    f.write(streamlit_app_code_fpdf2)

print("Streamlit app saved to 'app.py' successfully with fpdf2 compatibility.")
